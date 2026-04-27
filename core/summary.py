"""Generación de resúmenes y exportación de datos a CSV/Excel."""

import io
import pandas as pd
import re
from datetime import datetime
import json
import tempfile
from math import ceil
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU

try:
    import cairosvg
except Exception:  # pragma: no cover - entorno sin dependencia opcional
    cairosvg = None

try:
    from PIL import Image as PILImage
except Exception:  # pragma: no cover - entorno sin dependencia opcional
    PILImage = None

try:
    from PySide6.QtGui import QImage, QPainter
    from PySide6.QtSvg import QSvgRenderer
    from PySide6.QtWidgets import QApplication
except Exception:  # pragma: no cover - entorno sin dependencia opcional
    QImage = None
    QPainter = None
    QSvgRenderer = None
    QApplication = None


_QT_APP_REF = None

from core.model import (
    Piece,
    Project,
    build_piece_observations_display,
    normalize_piece_grain_direction,
    normalize_piece_observations,
)


def _excel_image_embedding_available() -> bool:
    """openpyxl requiere Pillow incluso para insertar PNG ya existentes."""

    return PILImage is not None


def export_summary(project: Project, output_csv: Path):
    """Exporta un resumen de piezas por módulo a un archivo CSV.
    
    Excluye piezas con espesor = 0.
    Módulos automáticos (escaneados) se muestran primero, luego los manuales.
    Columnas en orden: module; piece_id; piece_name; quantity; height; width; thickness; color; grain_direction; source
    """
    # Ordenar módulos: primero automáticos (is_manual=False), luego manuales (is_manual=True)
    modules_sorted = sorted(project.modules, key=lambda m: m.is_manual)
    
    rows = []
    for module in modules_sorted:
        module_quantity = _safe_int(getattr(module, "quantity", None), default=1)
        for piece in module.pieces:
            # Filtrar piezas con espesor 0
            if piece.thickness == 0 or piece.thickness is None:
                continue
            rows.append({
                "module": module.name,
                "piece_id": piece.id,
                "piece_name": piece.name or piece.id,
                "quantity": _effective_piece_quantity(piece.quantity, module_quantity),
                "height": piece.height,
                "width": piece.width,
                "thickness": piece.thickness,
                "color": piece.color,
                "grain_direction": normalize_piece_grain_direction(piece.grain_direction),
                "source": piece.cnc_source,
            })
    df = pd.DataFrame(rows, columns=["module", "piece_id", "piece_name", "quantity", "height", "width", "thickness", "color", "grain_direction", "source"])
    df.to_csv(output_csv, index=False, encoding="utf-8", sep=";")
    return df


def _is_valid_thickness(value) -> bool:
    if value is None:
        return False
    try:
        return float(value) > 0
    except (TypeError, ValueError):
        return False


def _safe_int(value, default=1) -> int:
    try:
        parsed = int(float(value))
        return parsed if parsed > 0 else default
    except (TypeError, ValueError):
        return default


def _effective_piece_quantity(piece_quantity, module_quantity) -> int:
    return _safe_int(piece_quantity, default=1) * _safe_int(module_quantity, default=1)


def _safe_float(value):
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _confirmed_dimension(value):
    parsed = _safe_float(value)
    if parsed is None:
        return None
    return int(parsed) if parsed.is_integer() else round(parsed, 2)


def _is_positive_dimension(value) -> bool:
    parsed = _safe_float(value)
    return parsed is not None and parsed > 0


def _piece_from_sheet_row(module_name: str, piece_row: dict) -> Piece:
    thickness = _safe_float(piece_row.get("thickness"))
    quantity = _safe_int(piece_row.get("quantity"), default=1)
    return Piece(
        id=str(piece_row.get("id") or piece_row.get("name") or "pieza").strip(),
        name=str(piece_row.get("name") or piece_row.get("id") or "pieza").strip(),
        quantity=quantity,
        width=_safe_float(piece_row.get("width")) or 0.0,
        height=_safe_float(piece_row.get("height")) or 0.0,
        thickness=thickness,
        color=piece_row.get("color"),
        grain_direction=normalize_piece_grain_direction(piece_row.get("grain_direction")),
        module_name=module_name,
        cnc_source=str(piece_row.get("source") or "").strip() or None,
        f6_source=str(piece_row.get("f6_source") or "").strip() or None,
        piece_type=piece_row.get("piece_type"),
        program_width=_safe_float(piece_row.get("program_width")),
        program_height=_safe_float(piece_row.get("program_height")),
        program_thickness=_safe_float(piece_row.get("program_thickness")),
    )


def _px_to_excel_width(px: int) -> float:
    """Convertir px a unidad de ancho de columna de Excel (aprox. Calibri 11)."""
    if px <= 12:
        return 1.0
    return round((px - 5) / 7, 2)


def _excel_width_for_text_values(
    values,
    *,
    minimum_width: float,
    maximum_width: float,
    padding_chars: float = 2.0,
) -> float:
    longest_line_length = 0
    for value in values:
        text = str(value or "").strip()
        if not text:
            continue
        longest_line_length = max(
            longest_line_length,
            max(len(line.strip()) for line in text.splitlines() if line.strip()) if text.splitlines() else len(text),
        )

    if longest_line_length <= 0:
        return minimum_width

    estimated_width = round((longest_line_length * 0.92) + padding_chars, 2)
    return max(minimum_width, min(maximum_width, estimated_width))


def _sanitize_filename(value: str) -> str:
    cleaned = re.sub(r"[^0-9A-Za-z._-]+", "_", str(value or "").strip())
    return cleaned.strip("._") or "pieza"


def _add_text_to_png(png_path: Path, text: str, font_size: int = 10) -> None:
    """Agrega texto centrado en la imagen PNG (sobre la imagen original)."""
    if PILImage is None:
        return
    
    try:
        from PIL import ImageDraw, ImageFont
        
        # Abrir imagen existente
        img = PILImage.open(png_path).convert("RGBA")
        img_w, img_h = img.size
        
        # Crear capa de texto semi-transparente
        txt_layer = PILImage.new("RGBA", (img_w, img_h), (255, 255, 255, 0))
        draw = ImageDraw.Draw(txt_layer)
        
        # Intentar usar una fuente específica, si no está disponible usar la default
        try:
            font = ImageFont.truetype("calibri.ttf", font_size)
        except Exception:
            try:
                font = ImageFont.truetype("arial.ttf", font_size)
            except Exception:
                font = ImageFont.load_default()
        
        # Calcular dimensiones del texto
        bbox = draw.textbbox((0, 0), text, font=font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]
        
        # Posición centrada en la imagen
        text_x = max(0, (img_w - text_width) // 2)
        text_y = max(0, (img_h - text_height) // 2)
        
        # Dibujar rectángulo de fondo semi-transparente
        padding = 5
        bg_left = max(0, text_x - padding)
        bg_top = max(0, text_y - padding)
        bg_right = min(img_w, text_x + text_width + padding)
        bg_bottom = min(img_h, text_y + text_height + padding)
        draw.rectangle(
            [(bg_left, bg_top), (bg_right, bg_bottom)],
            fill=(255, 255, 255, 200)  # Fondo blanco semi-transparente
        )
        
        # Dibujar texto en gris oscuro
        draw.text((text_x, text_y), text, fill=(50, 50, 50, 255), font=font)
        
        # Combinar capas
        img_with_text = PILImage.alpha_composite(img, txt_layer)
        
        # Guardar imagen modificada (convertir de vuelta a RGB para PNG)
        img_rgb = PILImage.new("RGB", img_with_text.size, (255, 255, 255))
        img_rgb.paste(img_with_text, mask=img_with_text.split()[3])
        img_rgb.save(png_path, format="PNG")
    except Exception:
        pass  # Si hay error, dejar la imagen como está


def _svg_to_png_for_excel(svg_path: Path, png_path: Path, max_width_px: int = 200) -> tuple[int, int] | None:
    """Convierte SVG a PNG para Excel y limita ancho sin perder aspecto."""

    if cairosvg is not None and PILImage is not None:
        try:
            cairosvg.svg2png(url=str(svg_path), write_to=str(png_path))
            with PILImage.open(png_path) as image:
                width_px, height_px = image.size
                if width_px <= 0 or height_px <= 0:
                    return None

                if width_px > max_width_px:
                    ratio = max_width_px / float(width_px)
                    resized = image.resize((max_width_px, max(1, int(round(height_px * ratio)))))
                    resized.save(png_path, format="PNG")
                    width_px, height_px = resized.size

            return width_px, height_px
        except Exception:
            pass

    # Fallback para entornos Windows donde CairoSVG puede no cargar librerías nativas.
    if QSvgRenderer is None or QImage is None or QPainter is None:
        return None

    # QSvgRenderer renderiza de forma estable cuando existe una app Qt activa.
    global _QT_APP_REF
    if QApplication is not None and QApplication.instance() is None:
        _QT_APP_REF = QApplication([])

    renderer = QSvgRenderer(str(svg_path))
    if not renderer.isValid():
        return None

    default_size = renderer.defaultSize()
    width_px = int(default_size.width()) if default_size.width() > 0 else max_width_px
    height_px = int(default_size.height()) if default_size.height() > 0 else max_width_px
    if width_px <= 0 or height_px <= 0:
        return None

    if width_px > max_width_px:
        ratio = max_width_px / float(width_px)
        width_px = max_width_px
        height_px = max(1, int(round(height_px * ratio)))

    image = QImage(width_px, height_px, QImage.Format_ARGB32)
    image.fill(0)
    painter = QPainter(image)
    renderer.render(painter)
    painter.end()

    if not image.save(str(png_path), "PNG"):
        return None

    return width_px, height_px


def _normalize_en_juego_sheet_cut_mode(value) -> str:
    raw = str(value or "").strip().lower()
    if raw in {"nesting", "corte nesting", "cut nesting"}:
        return "nesting"
    if raw in {"manual", "corte manual", "cut manual"}:
        return "manual"
    return "manual"


def _en_juego_sheet_replacement_enabled(config_data: dict, pieces: list[dict]) -> bool:
    settings = config_data.get("en_juego_settings")
    if not isinstance(settings, dict):
        return False
    if _normalize_en_juego_sheet_cut_mode(settings.get("cut_mode")) != "nesting":
        return False
    return any(bool(piece.get("en_juego", False)) and bool(piece.get("include_in_sheet", False)) for piece in pieces)


def _resolve_en_juego_output_path(
    project: Project,
    module_path: Path,
    module_name: str,
    config_data: dict,
) -> Path | None:
    raw_paths: list[str] = []
    for key in ("en_juego_output_path", "en_juego_pgmx_path"):
        raw_value = str(config_data.get(key) or "").strip()
        if raw_value:
            raw_paths.append(raw_value)

    synthesis_data = config_data.get("en_juego_synthesis")
    if isinstance(synthesis_data, dict):
        raw_value = str(synthesis_data.get("output_path") or "").strip()
        if raw_value:
            raw_paths.append(raw_value)

    bases = [module_path]
    project_root_value = str(getattr(project, "root_directory", "") or "").strip()
    if project_root_value:
        bases.append(Path(project_root_value))

    for raw_value in raw_paths:
        candidate = Path(raw_value)
        if candidate.is_file():
            return candidate
        if not candidate.is_absolute():
            for base in bases:
                based_candidate = base / candidate
                if based_candidate.is_file():
                    return based_candidate

    exact_candidates = [
        module_path / f"{module_name}_EnJuego.pgmx",
        module_path / f"{_sanitize_filename(module_name)}_EnJuego.pgmx",
    ]
    for candidate in exact_candidates:
        if candidate.is_file():
            return candidate

    discovered = sorted(
        (candidate for candidate in module_path.glob("*EnJuego.pgmx") if candidate.is_file()),
        key=lambda candidate: candidate.stat().st_mtime,
        reverse=True,
    )
    return discovered[0] if discovered else None


def _build_en_juego_sheet_svg(
    project: Project,
    module,
    module_path: Path,
    config_data: dict,
    temp_dir: Path,
) -> Path | None:
    en_juego_pgmx_path = _resolve_en_juego_output_path(project, module_path, module.name, config_data)
    if en_juego_pgmx_path is None:
        return None

    from core.pgmx_processing import build_piece_svg, parse_pgmx_for_piece

    en_juego_piece = Piece(
        id="EN_JUEGO",
        name="En-Juego",
        width=0.0,
        height=0.0,
        thickness=None,
        quantity=1,
        module_name=module.name,
        cnc_source=str(en_juego_pgmx_path),
    )
    drawing = parse_pgmx_for_piece(project, en_juego_piece, module_path)
    if drawing is None:
        return None

    svg_path = temp_dir / f"{_sanitize_filename(module.name)}_EnJuego.svg"
    try:
        build_piece_svg(en_juego_piece, drawing, svg_path)
    except Exception:
        return None
    return svg_path if svg_path.is_file() else None


def _load_module_sheet_data(
    project: Project,
    module,
    program_dimensions_cache: dict[tuple[str, str], tuple[float | None, float | None, float | None]],
) -> dict:
    from core.pgmx_processing import get_pgmx_program_dimension_notes

    config_path = Path(module.path) / "module_config.json"
    config_data = {}
    module_settings = {
        "herrajes_y_accesorios": "",
        "guias_y_bisagras": "",
        "detalles_de_obra": "",
    }

    if config_path.exists():
        try:
            config_data = json.loads(config_path.read_text(encoding="utf-8"))
        except Exception:
            config_data = {}
        module_settings.update(config_data.get("settings", {}))
        raw_pieces = config_data.get("pieces", [])
        pieces = [piece for piece in raw_pieces if _is_valid_thickness(piece.get("thickness"))]
        from core.model import PIECE_TYPE_ORDER as _PTO

        type_rank = {piece_type: index for index, piece_type in enumerate(_PTO)}
        pieces.sort(key=lambda piece: type_rank.get(piece.get("piece_type") or "", len(_PTO)))
    else:
        pieces = [
            {
                "id": piece.id,
                "name": piece.name or piece.id,
                "quantity": piece.quantity,
                "width": piece.width,
                "height": piece.height,
                "thickness": piece.thickness,
                "color": piece.color,
                "grain_direction": normalize_piece_grain_direction(piece.grain_direction),
                "source": piece.cnc_source,
                "f6_source": piece.f6_source,
                "program_width": piece.program_width,
                "program_height": piece.program_height,
                "program_thickness": piece.program_thickness,
                "include_in_sheet": False,
                "observations": "",
            }
            for piece in module.pieces
            if _is_valid_thickness(piece.thickness)
        ]

    piece_objects = [_piece_from_sheet_row(module.name, piece) for piece in pieces]
    program_notes = get_pgmx_program_dimension_notes(
        project,
        piece_objects,
        Path(module.path),
        cache=program_dimensions_cache,
    )
    for piece, program_note in zip(pieces, program_notes):
        piece["program_dimension_note"] = program_note
        piece["observations"] = normalize_piece_observations(piece.get("observations"))

    x_inferred, y_inferred, z_inferred = _derive_module_dimensions(module.name, pieces)
    x_val = _confirmed_dimension(module_settings.get("x")) or x_inferred
    y_val = _confirmed_dimension(module_settings.get("y")) or y_inferred
    z_val = _confirmed_dimension(module_settings.get("z")) or z_inferred
    return {
        "config_data": config_data,
        "module_settings": module_settings,
        "pieces": pieces,
        "dimensions": (x_val, y_val, z_val),
    }


def _prepare_module_sheet_images(
    project: Project,
    module,
    config_data: dict,
    pieces: list[dict],
    temp_dir: Path,
    *,
    max_width_px: int = 150,
) -> list[tuple[Path, int, int, str]]:
    prepared_images: list[tuple[Path, int, int, str]] = []
    if not _excel_image_embedding_available():
        return prepared_images

    en_juego_replaced_piece_ids: set[str] = set()
    if _en_juego_sheet_replacement_enabled(config_data, pieces):
        en_juego_svg_path = _build_en_juego_sheet_svg(
            project,
            module,
            Path(module.path),
            config_data,
            temp_dir,
        )
        if en_juego_svg_path is not None:
            en_juego_png_path = temp_dir / f"{_sanitize_filename(module.name)}_EnJuego.png"
            prepared_en_juego = _prepare_excel_drawing_image(
                en_juego_svg_path,
                en_juego_png_path,
                "En-Juego",
                max_width_px=max_width_px,
            )
            if prepared_en_juego is not None:
                prepared_images.append(prepared_en_juego)
                en_juego_replaced_piece_ids = {
                    str(piece.get("id") or "").strip()
                    for piece in pieces
                    if bool(piece.get("en_juego", False))
                }

    for idx, piece in enumerate(pieces):
        if not bool(piece.get("include_in_sheet", False)):
            continue
        if (
            bool(piece.get("en_juego", False))
            and str(piece.get("id") or "").strip() in en_juego_replaced_piece_ids
        ):
            continue

        piece_display_name = str(piece.get("name") or piece.get("id") or "pieza").strip()
        piece_slug = _sanitize_filename(piece_display_name)
        svg_path = Path(module.path) / f"{piece_slug}.svg"
        if not svg_path.is_file():
            continue

        png_path = temp_dir / f"{_sanitize_filename(module.name)}_{idx}_{piece_slug}.png"
        prepared_piece = _prepare_excel_drawing_image(
            svg_path,
            png_path,
            piece_display_name,
            max_width_px=max_width_px,
        )
        if prepared_piece is not None:
            prepared_images.append(prepared_piece)

    return prepared_images


def _prepare_excel_drawing_image(
    svg_path: Path,
    png_path: Path,
    display_name: str,
    *,
    max_width_px: int = 150,
) -> tuple[Path, int, int, str] | None:
    size = _svg_to_png_for_excel(svg_path, png_path, max_width_px=max_width_px)
    if not size:
        return None

    src_w, src_h = size
    if src_w <= 0 or src_h <= 0:
        return None

    if src_h > 150:
        ratio = 150.0 / float(src_h)
        target_w = int(round(src_w * ratio))
        target_h = 150
    else:
        target_w = src_w
        target_h = src_h

    _add_text_to_png(png_path, display_name, font_size=10)
    return png_path, target_w, target_h, display_name


def _pil_image_to_rgb_on_white(image):
    if image.mode in {"RGBA", "LA"} or "transparency" in image.info:
        rgba = image.convert("RGBA")
        background = PILImage.new("RGB", rgba.size, "white")
        background.paste(rgba, mask=rgba.split()[-1])
        return background
    return image.convert("RGB")


def _prepare_pdf_popup_drawing_image(
    svg_path: Path,
    png_path: Path,
    *,
    max_width_px: int = 1260,
    max_height_px: int = 900,
    width_scale: float = 3.0,
) -> tuple[Path, int, int] | None:
    size = _svg_to_png_for_excel(svg_path, png_path, max_width_px=max_width_px)
    if not size or PILImage is None:
        return None

    try:
        with PILImage.open(png_path) as image:
            rgb_image = _pil_image_to_rgb_on_white(image)
            source_w, source_h = rgb_image.size
            if source_w <= 0 or source_h <= 0:
                return None

            target_w = max(1, int(round(source_w * width_scale)))
            target_w = min(target_w, max_width_px)
            target_h = max(1, int(round(source_h * (target_w / float(source_w)))))
            if target_h > max_height_px:
                ratio = max_height_px / float(target_h)
                target_w = max(1, int(round(target_w * ratio)))
                target_h = max_height_px

            if (target_w, target_h) != rgb_image.size:
                resampling = getattr(getattr(PILImage, "Resampling", PILImage), "LANCZOS", 1)
                rgb_image = rgb_image.resize((target_w, target_h), resampling)
            rgb_image.save(png_path, format="PNG")
            width_px, height_px = rgb_image.size
    except Exception:
        return None

    return png_path, width_px, height_px


def _apply_outer_frame(ws, start_row: int, end_row: int, start_col: int = 1, end_col: int = 7):
    """Dibuja un marco exterior alrededor de un bloque rectangular."""

    if end_row < start_row:
        return

    side = Side(style="medium", color="FF808080")
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            left = side if col == start_col else cell.border.left
            right = side if col == end_col else cell.border.right
            top = side if row == start_row else cell.border.top
            bottom = side if row == end_row else cell.border.bottom
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def _extract_named_dimensions(module_name: str):
    """Extraer dimensiones desde nombre de módulo respetando convención ...-X-Z."""
    raw = module_name or ""
    cleaned = re.sub(r"^\s*mod\.?\s*\d+\s*-\s*", "", raw, flags=re.IGNORECASE)
    parts = [part.strip() for part in cleaned.split("-")]

    numeric_parts = []
    for part in parts:
        if re.fullmatch(r"\d+(?:[\.,]\d+)?", part):
            try:
                numeric_parts.append(float(part.replace(",", ".")))
            except ValueError:
                continue

    if len(numeric_parts) >= 2:
        return numeric_parts[-2], numeric_parts[-1]
    if len(numeric_parts) == 1:
        return numeric_parts[0], None
    return None, None


def _derive_module_dimensions(module_name: str, pieces: list[dict]):
    """Inferir X, Y, Z según: ancho total, altura total y profundidad sin frente."""
    x_named, z_named = _extract_named_dimensions(module_name)

    widths = []
    heights = []
    thicknesses = []
    lateral_heights = []
    span_heights = []

    for piece in pieces:
        piece_name = str(piece.get("name") or piece.get("id") or "").lower()
        width = _safe_float(piece.get("width"))
        height = _safe_float(piece.get("height"))
        thickness = _safe_float(piece.get("thickness"))

        if width is not None and width > 0:
            widths.append(width)
        if height is not None and height > 0:
            heights.append(height)
        if thickness is not None and thickness > 0:
            thicknesses.append(thickness)

        if "lateral" in piece_name and height is not None and height > 0:
            lateral_heights.append(height)

        if any(key in piece_name for key in ["fondo", "estante", "tapa", "puerta", "frente", "faja"]):
            if height is not None and height > 0:
                span_heights.append(height)

    max_thickness = max(thicknesses) if thicknesses else 0.0

    # Z: profundidad sin frente/tapas. En la mayoría de casos coincide con el mayor ancho de pieza.
    if z_named is not None:
        z_val = round(z_named, 2)
    elif widths:
        z_val = round(max(widths), 2)
    else:
        z_val = None

    # X: ancho total. Se prioriza nomenclatura del módulo (convención ...-X-Z).
    if x_named is not None:
        x_val = round(x_named, 2)
    else:
        # Fallback geométrico: usar piezas de "travesía" (fondo/estante/puerta/faja/tapa/frente).
        x_base = max(span_heights) if span_heights else (max(heights) if heights else None)
        x_val = round(x_base, 2) if x_base is not None else None

    # Y: altura total. Lateral + espesor superior/inferior cuando se puede inferir.
    if lateral_heights:
        y_base = max(lateral_heights)
    elif heights:
        y_base = max(heights)
    else:
        y_base = None

    y_val = round(y_base + max_thickness, 2) if y_base is not None else None

    # Normalizar enteros visuales (150.0 -> 150)
    def _compact_number(n):
        if n is None:
            return None
        return int(n) if abs(n - int(n)) < 1e-9 else n

    return _compact_number(x_val), _compact_number(y_val), _compact_number(z_val)


PDF_EXPORT_DPI = 150
PDF_PAGE_WIDTH = int(round(8.27 * PDF_EXPORT_DPI))
PDF_PAGE_HEIGHT = int(round(11.69 * PDF_EXPORT_DPI))
PDF_MARGIN_X = int(round(0.35 * PDF_EXPORT_DPI))
PDF_MARGIN_TOP = int(round(0.35 * PDF_EXPORT_DPI))
PDF_MARGIN_BOTTOM = int(round(0.3 * PDF_EXPORT_DPI))
PDF_HEADER_HEIGHT = int(round(1.05 * PDF_EXPORT_DPI))


def _load_pdf_font(size: int, *, bold: bool = False):
    from PIL import ImageFont

    font_names = (
        ("arialbd.ttf", "arial.ttf")
        if bold
        else ("arial.ttf", "calibri.ttf", "DejaVuSans.ttf")
    )
    for font_name in font_names:
        try:
            return ImageFont.truetype(font_name, size)
        except Exception:
            continue
    return ImageFont.load_default()


def _text_width(draw, text: str, font) -> int:
    bbox = draw.textbbox((0, 0), str(text or ""), font=font)
    return int(bbox[2] - bbox[0])


def _fit_text(draw, text, font, max_width: int) -> str:
    value = str(text or "")
    if _text_width(draw, value, font) <= max_width:
        return value
    ellipsis = "..."
    while value and _text_width(draw, f"{value}{ellipsis}", font) > max_width:
        value = value[:-1]
    return f"{value}{ellipsis}" if value else ellipsis


def _draw_fit_text(draw, xy: tuple[int, int], text, font, fill, max_width: int) -> None:
    draw.text(xy, _fit_text(draw, text, font, max_width), font=font, fill=fill)


def _pdf_column_metrics(content_width: int) -> tuple[list[int], list[int]]:
    base_widths = [110, 175, 45, 65, 65, 65, 95, 130]
    base_total = sum(base_widths)
    widths = [max(1, int(round(width * content_width / base_total))) for width in base_widths]
    widths[-1] += content_width - sum(widths)
    starts = [0]
    for width in widths[:-1]:
        starts.append(starts[-1] + width)
    return widths, starts


def _pdf_piece_rows(pieces: list[dict]) -> list[dict | None]:
    pieces_with_type = [piece for piece in pieces if piece.get("piece_type")]
    pieces_without_type = [piece for piece in pieces if not piece.get("piece_type")]
    rows: list[dict | None] = []
    last_h_index = None
    for index, piece in enumerate(pieces_with_type):
        if (piece.get("piece_type") or "").strip().upper() == "H":
            last_h_index = index

    for index, piece in enumerate(pieces_with_type):
        rows.append(piece)
        if index == last_h_index and (pieces_without_type or index < len(pieces_with_type) - 1):
            rows.append(None)

    if pieces_without_type and pieces_with_type:
        rows.append(None)
    rows.extend(pieces_without_type)
    return rows


PDF_CHECKBOX_SIZE = 16
PDF_CHECKBOX_GAP = 6
PDF_CHECKBOX_MARGIN_X = 6
PDF_CHECKBOX_MARGIN_Y = 7


def _pdf_checkboxes_per_line(column_width: int) -> int:
    available_width = max(1, column_width - (PDF_CHECKBOX_MARGIN_X * 2))
    return max(1, (available_width + PDF_CHECKBOX_GAP) // (PDF_CHECKBOX_SIZE + PDF_CHECKBOX_GAP))


def _pdf_piece_row_height(quantity: int, checkbox_column_width: int, base_row_height: int) -> int:
    boxes_per_line = _pdf_checkboxes_per_line(checkbox_column_width)
    lines = max(1, ceil(max(1, quantity) / boxes_per_line))
    checkbox_height = (
        PDF_CHECKBOX_MARGIN_Y * 2
        + (lines * PDF_CHECKBOX_SIZE)
        + ((lines - 1) * PDF_CHECKBOX_GAP)
    )
    return max(base_row_height, checkbox_height)


def _draw_pdf_checkboxes(
    draw,
    x: int,
    y: int,
    column_width: int,
    row_height: int,
    quantity: int,
) -> list[tuple[int, int]]:
    boxes_per_line = _pdf_checkboxes_per_line(column_width)
    quantity = max(1, int(quantity))
    positions: list[tuple[int, int]] = []
    for index in range(quantity):
        line_index = index // boxes_per_line
        col_index = index % boxes_per_line
        box_x = x + PDF_CHECKBOX_MARGIN_X + col_index * (PDF_CHECKBOX_SIZE + PDF_CHECKBOX_GAP)
        box_y = y + PDF_CHECKBOX_MARGIN_Y + line_index * (PDF_CHECKBOX_SIZE + PDF_CHECKBOX_GAP)
        if box_y + PDF_CHECKBOX_SIZE > y + row_height - 2:
            break
        positions.append((box_x, box_y))
        draw.rectangle(
            (box_x, box_y, box_x + PDF_CHECKBOX_SIZE, box_y + PDF_CHECKBOX_SIZE),
            outline="black",
            width=2,
        )
    return positions


def _prepare_pdf_piece_popups(
    project: Project,
    module,
    pieces: list[dict],
    temp_dir: Path,
    drawing_index_start: int,
) -> tuple[int, dict[str, dict]]:
    drawing_index = drawing_index_start
    popup_images: dict[str, dict] = {}
    module_path = Path(module.path)

    for piece in pieces:
        piece_display_name = str(piece.get("name") or piece.get("id") or "pieza").strip()
        piece_slug = _sanitize_filename(piece_display_name)
        svg_path = module_path / f"{piece_slug}.svg"
        if not svg_path.is_file():
            svg_path = temp_dir / f"{_sanitize_filename(module.name)}_{piece_slug}.svg"
            try:
                from core.pgmx_processing import build_piece_svg, parse_pgmx_for_piece

                piece_obj = _piece_from_sheet_row(module.name, piece)
                drawing_data = parse_pgmx_for_piece(project, piece_obj, module_path)
                if drawing_data is None:
                    continue
                build_piece_svg(piece_obj, drawing_data, svg_path)
            except Exception:
                continue
            if not svg_path.is_file():
                continue

        drawing_index += 1
        popup_key = f"img_{drawing_index:05d}"
        png_path = temp_dir / f"{popup_key}_{piece_slug}.png"
        prepared_popup = _prepare_pdf_popup_drawing_image(svg_path, png_path)
        if prepared_popup is None:
            continue

        prepared_path, width_px, height_px = prepared_popup
        piece["_pdf_popup_key"] = popup_key
        piece["_pdf_popup_width"] = width_px
        piece["_pdf_popup_height"] = height_px
        popup_images[popup_key] = {
            "path": prepared_path,
            "width": width_px,
            "height": height_px,
            "title": piece_display_name,
        }

    return drawing_index, popup_images


def _render_module_pdf_block(
    project: Project,
    module,
    module_data: dict,
) -> tuple["PILImage.Image", list[tuple[int, int]], list[dict]]:
    from PIL import Image, ImageDraw

    content_width = PDF_PAGE_WIDTH - (PDF_MARGIN_X * 2)
    col_widths, col_starts = _pdf_column_metrics(content_width)
    title_h = 38
    detail_h = 28
    row_h = 30
    gap_h = 18
    module_quantity = _safe_int(getattr(module, "quantity", None), default=1)
    module_settings = module_data["module_settings"]
    pieces = module_data["pieces"]
    x_val, y_val, z_val = module_data["dimensions"]

    detail_values = [
        ("herrajes_y_accesorios", (112, 48, 160)),
        ("guias_y_bisagras", (0, 112, 192)),
        ("detalles_de_obra", (84, 130, 53)),
    ]
    visible_details = [
        (str(module_settings.get(key) or "").strip(), color)
        for key, color in detail_values
        if str(module_settings.get(key) or "").strip()
    ]
    piece_rows = _pdf_piece_rows(pieces)
    piece_row_heights = [
        row_h if piece is None else _pdf_piece_row_height(
            _effective_piece_quantity(piece.get("quantity"), module_quantity),
            col_widths[0],
            row_h,
        )
        for piece in piece_rows
    ]
    block_height = (
        title_h
        + (len(visible_details) * detail_h)
        + gap_h
        + sum(piece_row_heights)
        + 10
    )
    block_height = max(block_height, title_h + 10)

    image = Image.new("RGB", (content_width, block_height), "white")
    draw = ImageDraw.Draw(image)
    title_font = _load_pdf_font(24, bold=True)
    dim_font = _load_pdf_font(20, bold=True)
    detail_font = _load_pdf_font(18, bold=True)
    row_font = _load_pdf_font(17)
    note_font = _load_pdf_font(15, bold=True)
    checkbox_positions: list[tuple[int, int]] = []
    piece_links: list[dict] = []

    y = 4
    _draw_fit_text(draw, (6, y + 5), module.name, title_font, "black", sum(col_widths[:3]) - 12)
    if all(_is_positive_dimension(value) for value in (x_val, y_val, z_val)):
        for label, value, col_index in (("X", x_val, 3), ("Y", y_val, 4), ("Z", z_val, 5)):
            _draw_fit_text(
                draw,
                (col_starts[col_index] + 4, y + 7),
                f"{label}: {value}",
                dim_font,
                "black",
                col_widths[col_index] - 8,
            )
    if module_quantity > 1:
        _draw_fit_text(
            draw,
            (col_starts[7] + 4, y + 9),
            f"{module_quantity} modulos.",
            note_font,
            (255, 0, 0),
            col_widths[7] - 8,
        )
    y += title_h

    for text, color in visible_details:
        _draw_fit_text(draw, (6, y + 4), text, detail_font, color, content_width - 12)
        y += detail_h

    y += gap_h

    for piece, current_row_h in zip(piece_rows, piece_row_heights):
        if piece is None:
            y += current_row_h
            continue
        effective_quantity = _effective_piece_quantity(piece.get("quantity"), module_quantity)
        checkbox_positions.extend(
            _draw_pdf_checkboxes(draw, col_starts[0], y, col_widths[0], current_row_h, effective_quantity)
        )
        popup_key = str(piece.get("_pdf_popup_key") or "").strip()
        note_value = build_piece_observations_display(
            piece.get("observations"),
            piece.get("program_dimension_note"),
        )
        values = [
            piece.get("name") or piece.get("id") or "",
            effective_quantity,
            _safe_float(piece.get("width")),
            _safe_float(piece.get("height")),
            _safe_float(piece.get("thickness")),
            piece.get("color") or "",
            note_value,
        ]
        for col_index, value in enumerate(values):
            target_col = col_index + 1
            font = note_font if target_col == 7 else row_font
            fill = (0, 82, 160) if target_col == 1 and popup_key else ((255, 0, 0) if target_col == 7 else "black")
            text_x = col_starts[target_col] + 4
            text_y = y + 6
            _draw_fit_text(
                draw,
                (text_x, text_y),
                "" if value is None else value,
                font,
                fill,
                col_widths[target_col] - 8,
            )
            if target_col == 1 and popup_key:
                fitted_name = _fit_text(draw, value, font, col_widths[target_col] - 8)
                underline_y = text_y + 20
                underline_w = min(_text_width(draw, fitted_name, font), col_widths[target_col] - 8)
                draw.line((text_x, underline_y, text_x + underline_w, underline_y), fill=(0, 82, 160), width=1)
                piece_links.append({
                    "key": popup_key,
                    "rect": (
                        col_starts[target_col] + 2,
                        y + 2,
                        col_widths[target_col] - 4,
                        min(current_row_h - 4, row_h),
                    ),
                })
        y += current_row_h

    draw.rectangle((0, 0, content_width - 1, min(block_height - 1, y + 4)), outline=(128, 128, 128), width=3)
    return image, checkbox_positions, piece_links


def _draw_production_pdf_header(draw, project: Project) -> int:
    title_font = _load_pdf_font(56)
    client_font = _load_pdf_font(24, bold=True)
    local_font = _load_pdf_font(36, bold=True)
    content_width = PDF_PAGE_WIDTH - (PDF_MARGIN_X * 2)
    x = PDF_MARGIN_X
    y = PDF_MARGIN_TOP
    _draw_fit_text(draw, (x, y), project.name or "Proyecto", title_font, "black", content_width)
    y += 70
    _draw_fit_text(draw, (x, y), f"Cliente: {project.client or '-'}", client_font, "black", content_width)
    y += 42
    _draw_fit_text(draw, (x, y), project.local or "", local_font, "black", content_width)
    return PDF_MARGIN_TOP + PDF_HEADER_HEIGHT


def _pdf_number(value: float) -> str:
    return f"{float(value):.4f}".rstrip("0").rstrip(".") or "0"


def _pdf_checkbox_appearance(size_pt: float, *, checked: bool) -> bytes:
    size = _pdf_number(size_pt)
    inset = _pdf_number(max(0.5, size_pt * 0.08))
    box_size = _pdf_number(max(1.0, size_pt - (2 * float(inset))))
    commands = [
        "q",
        "1 1 1 rg",
        f"0 0 {size} {size} re f",
        "0 0 0 RG",
        "1 w",
        f"{inset} {inset} {box_size} {box_size} re S",
    ]
    if checked:
        commands.extend([
            "2 w",
            f"{_pdf_number(size_pt * 0.22)} {_pdf_number(size_pt * 0.52)} m",
            f"{_pdf_number(size_pt * 0.42)} {_pdf_number(size_pt * 0.28)} l",
            f"{_pdf_number(size_pt * 0.78)} {_pdf_number(size_pt * 0.76)} l",
            "S",
        ])
    commands.append("Q")
    return "\n".join(commands).encode("ascii")


def _pdf_stream_object(dictionary: str, stream_data: bytes) -> bytes:
    return (
        f"<< {dictionary} /Length {len(stream_data)} >>\nstream\n".encode("ascii")
        + stream_data
        + b"\nendstream"
    )


def _pdf_literal_string(value: str) -> str:
    encoded = str(value or "").encode("cp1252", errors="replace").decode("cp1252")
    escaped = (
        encoded
        .replace("\\", "\\\\")
        .replace("(", "\\(")
        .replace(")", "\\)")
        .replace("\r", "\\r")
        .replace("\n", "\\n")
    )
    return f"({escaped})"


def _pdf_rect_from_pixels(
    x_px: float,
    y_px: float,
    width_px: float,
    height_px: float,
    page_height_px: int,
    scale: float,
) -> str:
    llx = x_px * scale
    lly = (page_height_px - (y_px + height_px)) * scale
    urx = (x_px + width_px) * scale
    ury = (page_height_px - y_px) * scale
    return " ".join(_pdf_number(value) for value in (llx, lly, urx, ury))


def _pdf_checkbox_rect(x_px: int, y_px: int, page_height_px: int, scale: float) -> str:
    return _pdf_rect_from_pixels(x_px, y_px, PDF_CHECKBOX_SIZE, PDF_CHECKBOX_SIZE, page_height_px, scale)


def _pdf_popup_rect_from_link(
    link_rect: tuple[float, float, float, float],
    popup_data: dict,
    page_width_px: int,
    page_height_px: int,
) -> tuple[int, int, int, int]:
    _link_x, link_y, _link_w, _link_h = link_rect
    image_width = int(popup_data.get("width") or 1)
    image_height = int(popup_data.get("height") or 1)
    popup_width = min(max(image_width + 36, 260), max(260, page_width_px - (PDF_MARGIN_X * 2) - 24))
    popup_height = min(max(image_height + 36, 200), max(200, page_height_px - (PDF_MARGIN_TOP + PDF_MARGIN_BOTTOM)))
    popup_x = max(PDF_MARGIN_X, page_width_px - PDF_MARGIN_X - popup_width - 12)
    popup_y = int(min(max(PDF_MARGIN_TOP, link_y - 16), max(PDF_MARGIN_TOP, page_height_px - PDF_MARGIN_BOTTOM - popup_height)))
    return popup_x, popup_y, popup_width, popup_height


def _pdf_popup_appearance(
    image_name: str,
    popup_width_pt: float,
    popup_height_pt: float,
    image_width_px: int,
    image_height_px: int,
    scale: float,
) -> bytes:
    margin = 8.0
    image_width_pt = max(1.0, image_width_px * scale)
    image_height_pt = max(1.0, image_height_px * scale)
    fit_ratio = min(
        (popup_width_pt - (margin * 2)) / image_width_pt,
        (popup_height_pt - (margin * 2)) / image_height_pt,
    )
    fitted_width = image_width_pt * fit_ratio
    fitted_height = image_height_pt * fit_ratio
    image_x = (popup_width_pt - fitted_width) / 2.0
    image_y = (popup_height_pt - fitted_height) / 2.0
    commands = [
        "q",
        "1 1 1 rg",
        f"0 0 {_pdf_number(popup_width_pt)} {_pdf_number(popup_height_pt)} re f",
        "0 0 0 RG",
        "1.2 w",
        f"0.6 0.6 {_pdf_number(popup_width_pt - 1.2)} {_pdf_number(popup_height_pt - 1.2)} re S",
        "q",
        f"{_pdf_number(fitted_width)} 0 0 {_pdf_number(fitted_height)} {_pdf_number(image_x)} {_pdf_number(image_y)} cm",
        f"/{image_name} Do",
        "Q",
        "Q",
    ]
    return "\n".join(commands).encode("ascii")


def _pdf_show_popup_javascript(active_key: str, popup_keys: list[str]) -> str:
    lines = ["try {"]
    for key in popup_keys:
        lines.append(f'var f_{key}=this.getField("{key}"); if (f_{key}) f_{key}.display = display.hidden;')
    lines.append(f'var active=this.getField("{active_key}"); if (active) active.display = display.visible;')
    lines.append("} catch (e) {}")
    return "\n".join(lines)


def _pdf_hide_popup_javascript(active_key: str) -> str:
    return f'try {{ var f=this.getField("{active_key}"); if (f) f.display = display.hidden; }} catch (e) {{}}'


def _pdf_checkbox_changed_javascript() -> str:
    return f'try {{ this.dirty = true; }} catch (e) {{}}'


def _write_interactive_pdf_page(
    page,
    output_pdf: Path,
    checkbox_positions: list[tuple[int, int]],
    popup_images: dict[str, dict],
    piece_links: list[dict],
) -> None:
    page_rgb = page.convert("RGB")
    image_buffer = io.BytesIO()
    page_rgb.save(image_buffer, format="JPEG", quality=95)
    image_data = image_buffer.getvalue()

    scale = 72.0 / float(PDF_EXPORT_DPI)
    page_width_pt = page_rgb.width * scale
    page_height_pt = page_rgb.height * scale
    content_stream = (
        f"q\n{_pdf_number(page_width_pt)} 0 0 {_pdf_number(page_height_pt)} 0 0 cm\n/Im0 Do\nQ\n"
    ).encode("ascii")

    objects: list[tuple[int, bytes]] = []
    field_refs: list[str] = []
    annot_refs: list[str] = []
    next_object_id = 6

    checkbox_object_ids: list[int] = []
    off_appearance_id = None
    yes_appearance_id = None
    if checkbox_positions:
        off_appearance_id = next_object_id
        yes_appearance_id = next_object_id + 1
        next_object_id += 2
        for _ in checkbox_positions:
            checkbox_object_ids.append(next_object_id)
            field_refs.append(f"{next_object_id} 0 R")
            annot_refs.append(f"{next_object_id} 0 R")
            next_object_id += 1

    popup_records: dict[str, dict] = {}
    visible_popup_keys = sorted({
        str(link.get("key") or "").strip()
        for link in piece_links
        if str(link.get("key") or "").strip() in popup_images
    })
    first_link_by_key: dict[str, tuple[float, float, float, float]] = {}
    for link in piece_links:
        key = str(link.get("key") or "").strip()
        if key and key not in first_link_by_key and isinstance(link.get("rect"), tuple):
            first_link_by_key[key] = link["rect"]

    for key in visible_popup_keys:
        popup_data = popup_images[key]
        popup_rect = _pdf_popup_rect_from_link(
            first_link_by_key[key],
            popup_data,
            page_rgb.width,
            page_rgb.height,
        )
        image_object_id = next_object_id
        appearance_object_id = next_object_id + 1
        widget_object_id = next_object_id + 2
        next_object_id += 3
        popup_records[key] = {
            "image_object_id": image_object_id,
            "appearance_object_id": appearance_object_id,
            "widget_object_id": widget_object_id,
            "rect": popup_rect,
            "data": popup_data,
        }
        field_refs.append(f"{widget_object_id} 0 R")
        annot_refs.append(f"{widget_object_id} 0 R")

    link_button_records: list[dict] = []
    for link in piece_links:
        key = str(link.get("key") or "").strip()
        rect = link.get("rect")
        if key not in popup_records or not isinstance(rect, tuple):
            continue
        object_id = next_object_id
        next_object_id += 1
        link_button_records.append({"object_id": object_id, "key": key, "rect": rect})
        field_refs.append(f"{object_id} 0 R")
        annot_refs.append(f"{object_id} 0 R")
    link_button_appearance_id = None
    if link_button_records:
        link_button_appearance_id = next_object_id
        next_object_id += 1

    acroform = ""
    if field_refs:
        acroform = f" /AcroForm << /Fields [{' '.join(field_refs)}] /NeedAppearances false >>"
    objects.append((1, f"<< /Type /Catalog /Pages 2 0 R{acroform} >>".encode("ascii")))
    objects.append((2, b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>"))

    annots = f" /Annots [{' '.join(annot_refs)}]" if annot_refs else ""
    page_body = (
        "<< /Type /Page /Parent 2 0 R "
        f"/MediaBox [0 0 {_pdf_number(page_width_pt)} {_pdf_number(page_height_pt)}] "
        "/Resources << /ProcSet [/PDF /ImageC] /XObject << /Im0 5 0 R >> >> "
        f"/Contents 4 0 R{annots} >>"
    )
    objects.append((3, page_body.encode("ascii")))
    objects.append((4, _pdf_stream_object("", content_stream)))
    image_dictionary = (
        f"/Type /XObject /Subtype /Image /Width {page_rgb.width} /Height {page_rgb.height} "
        "/ColorSpace /DeviceRGB /BitsPerComponent 8 /Filter /DCTDecode"
    )
    objects.append((5, _pdf_stream_object(image_dictionary, image_data)))

    if checkbox_positions and off_appearance_id is not None and yes_appearance_id is not None:
        checkbox_size_pt = PDF_CHECKBOX_SIZE * scale
        appearance_dictionary = (
            f"/Type /XObject /Subtype /Form /FormType 1 /BBox [0 0 "
            f"{_pdf_number(checkbox_size_pt)} {_pdf_number(checkbox_size_pt)}] /Resources << >>"
        )
        objects.append((
            off_appearance_id,
            _pdf_stream_object(appearance_dictionary, _pdf_checkbox_appearance(checkbox_size_pt, checked=False)),
        ))
        objects.append((
            yes_appearance_id,
            _pdf_stream_object(appearance_dictionary, _pdf_checkbox_appearance(checkbox_size_pt, checked=True)),
        ))
        for index, ((x_px, y_px), object_id) in enumerate(zip(checkbox_positions, checkbox_object_ids), start=1):
            rect = _pdf_checkbox_rect(x_px, y_px, page_rgb.height, scale)
            changed_js = _pdf_literal_string(_pdf_checkbox_changed_javascript())
            widget = (
                "<< /Type /Annot /Subtype /Widget /FT /Btn "
                f"/T (cb_{index:05d}) /F 4 /Ff 0 /Rect [{rect}] "
                "/V /Off /DV /Off /AS /Off "
                f"/AP << /N << /Off {off_appearance_id} 0 R /Yes {yes_appearance_id} 0 R >> >> "
                f"/AA << /U << /S /JavaScript /JS {changed_js} >> >> "
                "/MK << /BC [0 0 0] /BG [1 1 1] >> "
                "/BS << /W 1 /S /S >> /H /P /P 3 0 R >>"
            )
            objects.append((object_id, widget.encode("ascii")))

    for key, record in popup_records.items():
        popup_data = record["data"]
        popup_path = Path(popup_data["path"])
        try:
            with PILImage.open(popup_path) as popup_image:
                popup_rgb = _pil_image_to_rgb_on_white(popup_image)
                popup_buffer = io.BytesIO()
                popup_rgb.save(popup_buffer, format="JPEG", quality=95)
                popup_image_data = popup_buffer.getvalue()
                image_width = popup_rgb.width
                image_height = popup_rgb.height
        except Exception:
            popup_rgb = PILImage.new("RGB", (1, 1), "white")
            popup_buffer = io.BytesIO()
            popup_rgb.save(popup_buffer, format="JPEG", quality=95)
            popup_image_data = popup_buffer.getvalue()
            image_width = 1
            image_height = 1

        image_dictionary = (
            f"/Type /XObject /Subtype /Image /Width {image_width} /Height {image_height} "
            "/ColorSpace /DeviceRGB /BitsPerComponent 8 /Filter /DCTDecode"
        )
        image_object_id = int(record["image_object_id"])
        appearance_object_id = int(record["appearance_object_id"])
        widget_object_id = int(record["widget_object_id"])
        objects.append((image_object_id, _pdf_stream_object(image_dictionary, popup_image_data)))

        popup_x, popup_y, popup_w, popup_h = record["rect"]
        popup_w_pt = popup_w * scale
        popup_h_pt = popup_h * scale
        appearance_dictionary = (
            f"/Type /XObject /Subtype /Form /FormType 1 /BBox [0 0 {_pdf_number(popup_w_pt)} {_pdf_number(popup_h_pt)}] "
            f"/Resources << /XObject << /PopupImage {image_object_id} 0 R >> >>"
        )
        objects.append((
            appearance_object_id,
            _pdf_stream_object(
                appearance_dictionary,
                _pdf_popup_appearance("PopupImage", popup_w_pt, popup_h_pt, image_width, image_height, scale),
            ),
        ))

        popup_rect = _pdf_rect_from_pixels(popup_x, popup_y, popup_w, popup_h, page_rgb.height, scale)
        hide_js = _pdf_literal_string(_pdf_hide_popup_javascript(key))
        popup_widget = (
            "<< /Type /Annot /Subtype /Widget /FT /Btn "
            f"/T {_pdf_literal_string(key)} /F 6 /Ff 65536 /Rect [{popup_rect}] "
            f"/AP << /N {appearance_object_id} 0 R >> "
            f"/A << /S /JavaScript /JS {hide_js} >> "
            f"/AA << /U << /S /JavaScript /JS {hide_js} >> >> "
            "/MK << /BC [0 0 0] /BG [1 1 1] >> /BS << /W 1 /S /S >> /P 3 0 R >>"
        )
        objects.append((widget_object_id, popup_widget.encode("ascii")))

    if link_button_appearance_id is not None:
        objects.append((
            link_button_appearance_id,
            _pdf_stream_object(
                "/Type /XObject /Subtype /Form /FormType 1 /BBox [0 0 1 1] /Resources << >>",
                b"",
            ),
        ))

    for index, record in enumerate(link_button_records, start=1):
        key = record["key"]
        x_px, y_px, width_px, height_px = record["rect"]
        rect = _pdf_rect_from_pixels(x_px, y_px, width_px, height_px, page_rgb.height, scale)
        js = _pdf_literal_string(_pdf_show_popup_javascript(key, visible_popup_keys))
        link_button = (
            "<< /Type /Annot /Subtype /Widget /FT /Btn "
            f"/T {_pdf_literal_string(f'open_{index:05d}_{key}')} /F 4 /Ff 65536 /Rect [{rect}] "
            f"/AP << /N {link_button_appearance_id} 0 R >> "
            f"/A << /S /JavaScript /JS {js} >> "
            f"/AA << /U << /S /JavaScript /JS {js} >> >> "
            "/MK << >> /BS << /W 0 /S /S >> /H /P /P 3 0 R >>"
        )
        objects.append((int(record["object_id"]), link_button.encode("ascii")))

    max_object_id = max(object_id for object_id, _ in objects)
    objects.sort(key=lambda item: item[0])
    pdf_data = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0] * (max_object_id + 1)
    for object_id, body in objects:
        offsets[object_id] = len(pdf_data)
        pdf_data.extend(f"{object_id} 0 obj\n".encode("ascii"))
        pdf_data.extend(body)
        pdf_data.extend(b"\nendobj\n")

    xref_offset = len(pdf_data)
    pdf_data.extend(f"xref\n0 {max_object_id + 1}\n".encode("ascii"))
    pdf_data.extend(b"0000000000 65535 f \n")
    for object_id in range(1, max_object_id + 1):
        pdf_data.extend(f"{offsets[object_id]:010d} 00000 n \n".encode("ascii"))
    pdf_data.extend(
        f"trailer\n<< /Size {max_object_id + 1} /Root 1 0 R >>\n"
        f"startxref\n{xref_offset}\n%%EOF\n".encode("ascii")
    )
    output_pdf.write_bytes(pdf_data)


def export_production_sheet_pdf(project: Project, output_pdf: Path) -> Path:
    """Genera un PDF plano continuo de la planilla de produccion."""

    if PILImage is None:
        raise RuntimeError("Pillow es necesario para generar el PDF de planilla.")

    from PIL import Image, ImageDraw

    output_pdf = Path(output_pdf)
    program_dimensions_cache: dict[tuple[str, str], tuple[float | None, float | None, float | None]] = {}
    module_gap = 24
    blocks = []
    popup_images: dict[str, dict] = {}
    drawing_index = 0

    with tempfile.TemporaryDirectory() as temp_dir_name:
        temp_dir = Path(temp_dir_name)
        for module in project.modules:
            module_data = _load_module_sheet_data(project, module, program_dimensions_cache)
            drawing_index, module_popups = _prepare_pdf_piece_popups(
                project,
                module,
                module_data["pieces"],
                temp_dir,
                drawing_index,
            )
            popup_images.update(module_popups)
            blocks.append(_render_module_pdf_block(project, module, module_data))

        content_height = (
            PDF_MARGIN_TOP
            + PDF_HEADER_HEIGHT
            + sum(block.height for block, _checkboxes, _links in blocks)
            + (module_gap * max(0, len(blocks) - 1))
            + PDF_MARGIN_BOTTOM
        )
        page_height = max(PDF_PAGE_HEIGHT, content_height)
        page = Image.new("RGB", (PDF_PAGE_WIDTH, page_height), "white")
        draw = ImageDraw.Draw(page)
        current_y = _draw_production_pdf_header(draw, project)
        checkbox_positions: list[tuple[int, int]] = []
        piece_links: list[dict] = []

        for index, (block, block_checkboxes, block_links) in enumerate(blocks):
            page.paste(block, (PDF_MARGIN_X, current_y))
            checkbox_positions.extend(
                (PDF_MARGIN_X + x, current_y + y)
                for x, y in block_checkboxes
            )
            for link in block_links:
                x, y, width, height = link["rect"]
                piece_links.append({
                    "key": link["key"],
                    "rect": (PDF_MARGIN_X + x, current_y + y, width, height),
                })
            current_y += block.height
            if index < len(blocks) - 1:
                current_y += module_gap

        output_pdf.parent.mkdir(parents=True, exist_ok=True)
        _write_interactive_pdf_page(page, output_pdf, checkbox_positions, popup_images, piece_links)
    return output_pdf


def export_production_sheet(project: Project, output_xlsx: Path):
    """Generar planilla Excel de producción sin plantilla, desde datos del sistema."""
    from core.pgmx_processing import get_pgmx_program_dimension_notes

    wb = Workbook()
    ws = wb.active
    ws.title = "Planilla"
    program_dimensions_cache: dict[tuple[str, str], tuple[float | None, float | None, float | None]] = {}
    excel_images_supported = _excel_image_embedding_available()
    observations_min_width = _px_to_excel_width(120)
    observations_max_width = _px_to_excel_width(280)
    observations_font = Font(name="Calibri", size=9, bold=True, color="FFFF0000")
    observations_values: list[str] = []

    # Configurar formato de impresión
    # Tamaño A4 (paperSize=9)
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "portrait"
    
    # Escala: Ajustar al ancho de todas las columnas (ancho = 1)
    ws.page_setup.fitToHeight = 0  # Sin límite de altura
    ws.page_setup.fitToWidth = 1   # Ajustar a 1 página de ancho
    ws.print_options.horizontalCentered = True
    
    # Márgenes (en pulgadas): 1.5cm superior, 0.5cm otros
    ws.page_margins.left = 0.19685    # 0.5 cm en pulgadas
    ws.page_margins.right = 0.19685   # 0.5 cm en pulgadas
    ws.page_margins.top = 0.59055     # 1.5 cm en pulgadas
    ws.page_margins.bottom = 0.19685  # 0.5 cm en pulgadas
    ws.page_margins.header = 0.0
    ws.page_margins.footer = 0.0

    # Anchos de columnas solicitados (en px), convertidos a unidad Excel.
    ws.column_dimensions["A"].width = _px_to_excel_width(200)
    ws.column_dimensions["B"].width = _px_to_excel_width(50)
    ws.column_dimensions["C"].width = _px_to_excel_width(70)
    ws.column_dimensions["D"].width = _px_to_excel_width(70)
    ws.column_dimensions["E"].width = _px_to_excel_width(70)
    ws.column_dimensions["F"].width = _px_to_excel_width(70)
    ws.column_dimensions["G"].width = _px_to_excel_width(120)

    col_widths_px = [200, 50, 70, 70, 70, 70, 120]
    block_width_px = sum(col_widths_px)

    def marker_for_x(row_1_based: int, x_offset_px: float) -> AnchorMarker:
        """Convierte un offset X (px) dentro de A:G a (columna, offset local)."""

        x = max(0.0, min(float(x_offset_px), float(block_width_px - 1)))
        remaining = x
        col_idx = 0
        for idx, width in enumerate(col_widths_px):
            if remaining < width or idx == len(col_widths_px) - 1:
                col_idx = idx
                break
            remaining -= width

        return AnchorMarker(
            col=col_idx,
            row=row_1_based - 1,
            colOff=pixels_to_EMU(int(round(remaining))),
            rowOff=0,
        )

    # Encabezado principal
    ws["A1"] = project.name or "Proyecto"
    ws["A1"].font = Font(name="Calibri", size=50, bold=False)

    ws["A2"] = f"Cliente: {project.client or '-'}"
    ws["A2"].font = Font(name="Calibri", size=14, bold=True)

    # Local es opcional
    ws["A4"] = project.local or ""
    ws["A4"].font = Font(name="Calibri", size=24, bold=True)

    # Alturas base de encabezado.
    ws.row_dimensions[1].height = 64.5
    ws.row_dimensions[2].height = 18.75
    ws.row_dimensions[4].height = 31.5

    with tempfile.TemporaryDirectory() as temp_dir_name:
        temp_dir = Path(temp_dir_name)

        current_row = 6
        for module in project.modules:
            module_quantity = _safe_int(getattr(module, "quantity", None), default=1)
            config_path = Path(module.path) / "module_config.json"
            config_data = {}
            module_settings = {
                "herrajes_y_accesorios": "",
                "guias_y_bisagras": "",
                "detalles_de_obra": "",
            }

            if config_path.exists():
                try:
                    config_data = json.loads(config_path.read_text(encoding="utf-8"))
                except Exception:
                    config_data = {}
                module_settings.update(config_data.get("settings", {}))
                raw_pieces = config_data.get("pieces", [])
                pieces = [piece for piece in raw_pieces if _is_valid_thickness(piece.get("thickness"))]
                from core.model import PIECE_TYPE_ORDER as _PTO
                _type_rank = {t: i for i, t in enumerate(_PTO)}
                pieces.sort(key=lambda p: _type_rank.get(p.get("piece_type") or "", len(_PTO)))
            else:
                pieces = [
                    {
                        "id": piece.id,
                        "name": piece.name or piece.id,
                        "quantity": piece.quantity,
                        "width": piece.width,
                        "height": piece.height,
                        "thickness": piece.thickness,
                        "color": piece.color,
                        "grain_direction": normalize_piece_grain_direction(piece.grain_direction),
                        "source": piece.cnc_source,
                        "f6_source": piece.f6_source,
                        "program_width": piece.program_width,
                        "program_height": piece.program_height,
                        "program_thickness": piece.program_thickness,
                        "include_in_sheet": False,
                        "observations": "",
                    }
                    for piece in module.pieces
                    if _is_valid_thickness(piece.thickness)
                ]

            piece_objects = [_piece_from_sheet_row(module.name, piece) for piece in pieces]
            program_notes = get_pgmx_program_dimension_notes(
                project,
                piece_objects,
                Path(module.path),
                cache=program_dimensions_cache,
            )
            for piece, program_note in zip(pieces, program_notes):
                piece["program_dimension_note"] = program_note
                piece["observations"] = normalize_piece_observations(piece.get("observations"))

            x_inferred, y_inferred, z_inferred = _derive_module_dimensions(module.name, pieces)
            x_val = _confirmed_dimension(module_settings.get("x")) or x_inferred
            y_val = _confirmed_dimension(module_settings.get("y")) or y_inferred
            z_val = _confirmed_dimension(module_settings.get("z")) or z_inferred

            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
            module_title_cell = ws.cell(row=current_row, column=1, value=module.name)
            module_title_cell.font = Font(name="Calibri", size=14, bold=True)
            ws.row_dimensions[current_row].height = 18.75
            module_observation = ""
            if module_quantity > 1:
                module_observation = f"{module_quantity} módulos."
                module_observation_cell = ws.cell(row=current_row, column=7, value=module_observation)
                module_observation_cell.font = observations_font
                observations_values.append(module_observation)
            has_dimensions = all([
                _is_positive_dimension(x_val),
                _is_positive_dimension(y_val),
                _is_positive_dimension(z_val),
            ])

            dim_font = Font(name="Calibri", size=12, bold=True)
            if has_dimensions:
                dim_x_cell = ws.cell(row=current_row, column=4, value=f"X: {x_val}")
                dim_y_cell = ws.cell(row=current_row, column=5, value=f"Y: {y_val}")
                dim_z_cell = ws.cell(row=current_row, column=6, value=f"Z: {z_val}")
                dim_x_cell.font = dim_font
                dim_y_cell.font = dim_font
                dim_z_cell.font = dim_font
            else:
                ws.cell(row=current_row, column=4, value=None)
                ws.cell(row=current_row, column=5, value=None)
                ws.cell(row=current_row, column=6, value=None)

            detail_row = current_row + 1
            herrajes_value = str(module_settings.get("herrajes_y_accesorios") or "").strip()
            if herrajes_value:
                herrajes_cell = ws.cell(row=detail_row, column=1, value=herrajes_value)
                herrajes_cell.font = Font(name="Calibri", size=11, bold=True, color="FF7030A0")
                detail_row += 1

            guias_value = str(module_settings.get("guias_y_bisagras") or "").strip()
            if guias_value:
                guias_cell = ws.cell(row=detail_row, column=1, value=guias_value)
                guias_cell.font = Font(name="Calibri", size=11, bold=True, color="FF0070C0")
                detail_row += 1

            detalles_value = str(module_settings.get("detalles_de_obra") or "").strip()
            if detalles_value:
                detalles_cell = ws.cell(row=detail_row, column=1, value=detalles_value)
                detalles_cell.font = Font(name="Calibri", size=11, bold=True, color="FF548235")
                detail_row += 1

            # Leave exactly one blank row after the last detail/settings row.
            images_start_row = detail_row + 1
            images_end_row = images_start_row - 1  # No images initially

            # Insert drawings immediately after details: up to 3 per row, max height 4cm (150px).
            prepared_images: list[tuple[Path, int, int, str]] = []
            if excel_images_supported:
                en_juego_replaced_piece_ids: set[str] = set()
                if _en_juego_sheet_replacement_enabled(config_data, pieces):
                    en_juego_svg_path = _build_en_juego_sheet_svg(
                        project,
                        module,
                        Path(module.path),
                        config_data,
                        temp_dir,
                    )
                    if en_juego_svg_path is not None:
                        en_juego_png_path = temp_dir / f"{_sanitize_filename(module.name)}_EnJuego.png"
                        prepared_en_juego = _prepare_excel_drawing_image(
                            en_juego_svg_path,
                            en_juego_png_path,
                            "En-Juego",
                            max_width_px=150,
                        )
                        if prepared_en_juego is not None:
                            prepared_images.append(prepared_en_juego)
                            en_juego_replaced_piece_ids = {
                                str(piece.get("id") or "").strip()
                                for piece in pieces
                                if bool(piece.get("en_juego", False))
                            }

                for idx, piece in enumerate(pieces):
                    if not bool(piece.get("include_in_sheet", False)):
                        continue
                    if (
                        bool(piece.get("en_juego", False))
                        and str(piece.get("id") or "").strip() in en_juego_replaced_piece_ids
                    ):
                        continue

                    piece_display_name = str(piece.get("name") or piece.get("id") or "pieza").strip()
                    piece_slug = _sanitize_filename(piece_display_name)
                    svg_path = Path(module.path) / f"{piece_slug}.svg"
                    if not svg_path.is_file():
                        continue

                    # Max 4cm height (150px at 96 DPI)
                    png_path = temp_dir / f"{_sanitize_filename(module.name)}_{idx}_{piece_slug}.png"
                    prepared_piece = _prepare_excel_drawing_image(
                        svg_path,
                        png_path,
                        piece_display_name,
                        max_width_px=150,
                    )
                    if prepared_piece is not None:
                        prepared_images.append(prepared_piece)

            # Insertar imágenes: 3 por fila
            if prepared_images:
                drawing_anchor_row = images_start_row
                for chunk_start in range(0, len(prepared_images), 3):
                    chunk = prepared_images[chunk_start:chunk_start + 3]
                    count = len(chunk)
                    if count == 0:
                        continue

                    # Espaciado para 3 imágenes distribuidas en las 7 columnas:
                    # Imagen 1: columnas A:B (ancho 250px)
                    # Imagen 2: columnas C:D (ancho 140px)
                    # Imagen 3: columnas E:F:G (ancho 260px)
                    block_configs = [
                        {"start_col": 0, "end_col": 2, "width": sum(col_widths_px[:2])},
                        {"start_col": 2, "end_col": 4, "width": sum(col_widths_px[2:4])},
                        {"start_col": 4, "end_col": 7, "width": sum(col_widths_px[4:])},
                    ]
                    
                    max_h = 0
                    for idx_in_row, (png_path, img_w, img_h, piece_name) in enumerate(chunk):
                        if idx_in_row >= len(block_configs):
                            break
                            
                        config = block_configs[idx_in_row]
                        block_w = config["width"]
                        col_offset_start = sum(col_widths_px[:config["start_col"]])
                        
                        # Centrar imagen en su bloque asignado
                        centered_x = col_offset_start + max(0.0, (block_w - img_w) / 2.0)
                        
                        try:
                            image = XLImage(str(png_path))
                        except Exception:
                            continue
                        image.width = int(img_w)
                        image.height = int(img_h)

                        x_offset_px = int(round(centered_x))
                        image.anchor = OneCellAnchor(
                            _from=marker_for_x(drawing_anchor_row, x_offset_px),
                            ext=XDRPositiveSize2D(
                                pixels_to_EMU(int(image.width)),
                                pixels_to_EMU(int(image.height)),
                            ),
                        )
                        ws.add_image(image)
                        max_h = max(max_h, img_h)

                    ws.row_dimensions[drawing_anchor_row].height = max(
                        ws.row_dimensions[drawing_anchor_row].height or 15.0,
                        round(max_h * 0.75, 2),
                    )
                    images_end_row = drawing_anchor_row
                    drawing_anchor_row += 1

            # Piezas empiezan después de las imágenes
            pieces_start = images_end_row + 2 if prepared_images else images_start_row

            # Separar piezas en dos grupos: con piece_type y sin piece_type
            pieces_with_type = [p for p in pieces if p.get("piece_type")]
            pieces_without_type = [p for p in pieces if not p.get("piece_type")]

            # Encontrar índice de la última pieza tipo "H"
            last_h_index = None
            for i, piece in enumerate(pieces_with_type):
                if (piece.get("piece_type") or "").strip().upper() == "H":
                    last_h_index = i

            extra_rows = 0  # blank rows inserted so far (shifts subsequent offsets)
            
            # Procesar piezas con piece_type
            for offset, piece in enumerate(pieces_with_type):
                row = pieces_start + offset + extra_rows
                ws.cell(row=row, column=1, value=piece.get("name") or piece.get("id") or "")
                ws.cell(
                    row=row,
                    column=2,
                    value=_effective_piece_quantity(piece.get("quantity"), module_quantity),
                )
                ws.cell(row=row, column=3, value=_safe_float(piece.get("width")))
                ws.cell(row=row, column=4, value=_safe_float(piece.get("height")))
                ws.cell(row=row, column=5, value=_safe_float(piece.get("thickness")))
                ws.cell(row=row, column=6, value=piece.get("color") or "")
                note_value = build_piece_observations_display(
                    piece.get("observations"),
                    piece.get("program_dimension_note"),
                )
                col_g_cell = ws.cell(row=row, column=7, value=note_value)
                col_g_cell.font = observations_font
                observations_values.append(str(note_value))

                # Insertar fila vacía después de la última pieza tipo "H" si hay más piezas
                if offset == last_h_index and (pieces_without_type or offset < len(pieces_with_type) - 1):
                    extra_rows += 1

            # Agregar fila vacía si hay piezas sin piece_type
            if pieces_without_type and pieces_with_type:
                extra_rows += 1
            
            # Procesar piezas sin piece_type (agregadas manualmente)
            for offset, piece in enumerate(pieces_without_type):
                row = pieces_start + len(pieces_with_type) + offset + extra_rows
                ws.cell(row=row, column=1, value=piece.get("name") or piece.get("id") or "")
                ws.cell(
                    row=row,
                    column=2,
                    value=_effective_piece_quantity(piece.get("quantity"), module_quantity),
                )
                ws.cell(row=row, column=3, value=_safe_float(piece.get("width")))
                ws.cell(row=row, column=4, value=_safe_float(piece.get("height")))
                ws.cell(row=row, column=5, value=_safe_float(piece.get("thickness")))
                ws.cell(row=row, column=6, value=piece.get("color") or "")
                note_value = build_piece_observations_display(
                    piece.get("observations"),
                    piece.get("program_dimension_note"),
                )
                col_g_cell = ws.cell(row=row, column=7, value=note_value)
                col_g_cell.font = observations_font
                observations_values.append(str(note_value))

            module_last_piece_row = pieces_start + len(pieces_with_type) + len(pieces_without_type) + extra_rows - 1
            content_end_row = max(module_last_piece_row, images_end_row)
            _apply_outer_frame(ws, current_row, content_end_row, start_col=1, end_col=7)

            # Leave two blank rows between framed module blocks.
            current_row = content_end_row + 3

        ws.column_dimensions["G"].width = _excel_width_for_text_values(
            observations_values,
            minimum_width=observations_min_width,
            maximum_width=observations_max_width,
        )

        output_xlsx.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_xlsx)
    return output_xlsx
